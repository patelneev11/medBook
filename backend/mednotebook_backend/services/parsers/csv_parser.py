import csv
import io
import logging
import re
from datetime import datetime, timedelta
from typing import Optional

import openpyxl
import pandas as pd

from .encoding import decode_bytes
from .exceptions import ParserException
from .tabular import DATE_COLUMN_RE, LAB_UNIT_NAME_RE, LAB_VALUE_RE, PATIENT_ID_COLUMN_RE, format_table_as_markdown

logger = logging.getLogger("mednotebook.parsers.csv")

_FULL_TABLE_ROW_LIMIT = 500
_LARGE_DATASET_PREVIEW_ROWS = 100
_NUMERIC_RE = re.compile(r"^[+-]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?%?$")
_EXCEL_EPOCH = datetime(1899, 12, 30)


# ── CSV ──────────────────────────────────────────────────────────────────────

def extract_text_from_csv(file_content: bytes, filename: str) -> dict:
    """Extract a text report (summary + table) from a CSV's raw bytes.

    Handles missing/duplicated header rows, comment lines, mixed dtypes,
    and flags likely medical columns (patient IDs, dates, lab values) in
    the returned metadata.
    """
    if not file_content:
        raise ParserException("Empty file content", file_type="csv")

    text = decode_bytes(file_content, file_type="csv")
    delimiter = _sniff_delimiter(text)
    grid = _text_to_grid(text, delimiter)

    if not grid:
        raise ParserException("No data found in CSV file", file_type="csv")

    header_rows = _detect_header_rows(grid)
    df = _grid_to_dataframe(grid, header_rows)

    report_text, stats = _build_dataframe_report(df, label=filename)
    word_count = len(report_text.split())

    return {
        "text": report_text,
        "pages": [{"page_number": 1, "text": report_text, "word_count": word_count}],
        "page_count": 1,
        "word_count": word_count,
        "is_scanned": False,
        "extraction_method": "pandas",
        "metadata": {
            "title": filename,
            "author": None,
            "created_date": None,
            "modified_date": None,
            **stats,
        },
    }


def _sniff_delimiter(text: str) -> str:
    sample_lines = [ln for ln in text.splitlines() if not ln.lstrip().startswith("#")][:10]
    sample = "\n".join(sample_lines)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        return ","


def _text_to_grid(text: str, delimiter: str) -> list:
    lines = [ln for ln in text.splitlines() if not ln.lstrip().startswith("#")]
    reader = csv.reader(lines, delimiter=delimiter)
    return [row for row in reader if any(cell.strip() for cell in row)]


# ── Excel ────────────────────────────────────────────────────────────────────

def extract_text_from_excel(file_content: bytes, filename: str) -> dict:
    """Extract a text report from an Excel workbook's raw bytes, one "page"
    per visible sheet, processed with the same logic as extract_text_from_csv.
    """
    if not file_content:
        raise ParserException("Empty file content", file_type="excel")

    try:
        # data_only=True: read formulas' cached calculated values, not the
        # formula text. Only works if the file was last saved by an app that
        # cached results (real Excel does; a workbook never opened in Excel
        # may have no cached value and read as None).
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as exc:
        raise ParserException("Could not open Excel file", file_type="excel", original_error=exc) from exc

    pages = []
    sheet_names = []
    sheet_row_counts = {}
    medical = {"patient_id_columns": set(), "date_columns": set(), "lab_value_columns": set()}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue  # skip hidden / veryHidden sheets

        grid, captions = _sheet_to_grid(ws)
        if not grid:
            continue

        header_rows = _detect_header_rows(grid)
        df = _grid_to_dataframe(grid, header_rows)
        df = _fix_date_serials(df)

        report_text, stats = _build_dataframe_report(df, label=sheet_name)
        if captions:
            report_text = "\n".join(captions) + "\n\n" + report_text

        sheet_names.append(sheet_name)
        sheet_row_counts[sheet_name] = stats["row_count"]
        for key in medical:
            medical[key].update(stats["medical_columns"][key])

        page_text = f"=== Sheet: {sheet_name} ===\n\n{report_text}"
        pages.append({
            "page_number": len(pages) + 1,
            "text": page_text,
            "word_count": len(page_text.split()),
        })

    if not pages:
        raise ParserException("No visible sheets with data found in workbook", file_type="excel")

    full_text = "\n\n".join(p["text"] for p in pages)
    word_count = sum(p["word_count"] for p in pages)
    props = wb.properties

    return {
        "text": full_text,
        "pages": pages,
        "page_count": len(pages),
        "word_count": word_count,
        "is_scanned": False,
        "extraction_method": "openpyxl",
        "metadata": {
            "title": props.title or filename,
            "author": props.creator or None,
            "created_date": props.created.isoformat() if props.created else None,
            "modified_date": props.modified.isoformat() if props.modified else None,
            "sheet_names": sheet_names,
            "sheet_row_counts": sheet_row_counts,
            "medical_columns": {k: sorted(v) for k, v in medical.items()},
        },
    }


def _sheet_to_grid(ws) -> tuple:
    if ws.max_row is None or ws.max_column is None:
        return [], []

    grid = [[cell.value for cell in row] for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column,
    )]

    # openpyxl only stores a value on the top-left cell of a merged range —
    # every other cell in the range reads back as None. Fill the whole range
    # with the top-left value ("unmerge and fill down").
    for merged_range in ws.merged_cells.ranges:
        top_value = grid[merged_range.min_row - 1][merged_range.min_col - 1]
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                grid[r - 1][c - 1] = top_value

    grid = [row for row in grid if any(cell is not None and str(cell).strip() != "" for cell in row)]

    # A row that's a single merged cell spanning (nearly) the whole width
    # reads, after fill-down, as the identical value repeated across every
    # column — a section banner/title, not a header or data row. Pull those
    # out as standalone captions instead of feeding them into header
    # detection, where they'd get flattened into every column name.
    captions = []
    data_grid = []
    for row in grid:
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        distinct = {str(c).strip() for c in non_empty}
        if len(non_empty) >= 2 and len(distinct) == 1 and len(non_empty) >= 0.6 * len(row):
            captions.append(str(non_empty[0]).strip())
        else:
            data_grid.append(row)

    return data_grid, captions


def _fix_date_serials(df: pd.DataFrame) -> pd.DataFrame:
    """openpyxl already returns real datetime objects for properly
    date-formatted cells in most cases. This is a defensive fallback for
    date-named columns that still came through as raw Excel serial numbers
    (e.g. an unusual number_format openpyxl didn't recognize as a date).
    """
    for col in df.columns:
        if not DATE_COLUMN_RE.search(str(col)):
            continue
        df[col] = df[col].apply(_maybe_convert_excel_serial)
    return df


def _maybe_convert_excel_serial(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and 1 <= value <= 100_000:
        try:
            return _EXCEL_EPOCH + timedelta(days=float(value))
        except (OverflowError, ValueError):
            return value
    return value


# ── Shared grid / header detection (CSV text rows or Excel sheet cells) ──────

def _looks_numeric(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float, datetime)):
        return True
    s = str(value).strip()
    if not s:
        return False
    return bool(_NUMERIC_RE.match(s))


def _column_is_numeric_like(values: list) -> bool:
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    if not vals:
        return False
    return sum(1 for v in vals if _looks_numeric(v)) / len(vals) >= 0.7


def _row_looks_like_header(row: list, data_rows: list, default_when_ambiguous: bool) -> bool:
    """A row looks like a header if it's non-numeric labels sitting above
    columns whose actual data is numeric — the classic header/data contrast,
    checked per-column rather than by row-wide ratios (a row with mostly
    text columns, e.g. patient IDs + notes, would otherwise dilute a real
    numeric-column contrast signal into a "looks like data" false negative).

    `default_when_ambiguous` controls the tie-break when a column gives no
    signal either way (e.g. an all-text dataset with no numeric columns at
    all). It should be True when checking for the presence of a header at
    all (headers are the common case) and False when checking for a SECOND
    header row (wrongly claiming a real data row as a header silently drops
    that row's data — worse than the reverse mistake).
    """
    non_empty_idx = [i for i, c in enumerate(row) if c is not None and str(c).strip() != ""]
    if not non_empty_idx:
        return False

    # A real header label is very rarely itself a bare number.
    numeric_header_cells = sum(1 for i in non_empty_idx if _looks_numeric(row[i]))
    if numeric_header_cells / len(non_empty_idx) > 0.3:
        return False

    if not data_rows:
        return default_when_ambiguous

    header_evidence = 0  # column's data is numeric, this row's value isn't
    data_evidence = 0    # column's data is numeric, this row's value is too
    for i in non_empty_idx:
        col_values = [r[i] for r in data_rows if i < len(r)]
        if not col_values or not _column_is_numeric_like(col_values):
            continue  # text column — uninformative, no signal either way
        if _looks_numeric(row[i]):
            data_evidence += 1
        else:
            header_evidence += 1

    if header_evidence == 0 and data_evidence == 0:
        return default_when_ambiguous
    return header_evidence >= data_evidence if default_when_ambiguous else header_evidence > data_evidence


def _detect_header_rows(grid: list) -> int:
    if not grid:
        return 0
    if not _row_looks_like_header(grid[0], grid[1:11], default_when_ambiguous=True):
        return 0
    if len(grid) > 1 and _row_looks_like_header(grid[1], grid[2:11], default_when_ambiguous=False):
        return 2
    return 1


def _grid_to_dataframe(grid: list, header_rows: int) -> pd.DataFrame:
    if not grid:
        return pd.DataFrame()

    max_cols = max(len(row) for row in grid)
    padded = [list(row) + [None] * (max_cols - len(row)) for row in grid]

    if header_rows == 0:
        columns = [f"Column_{i + 1}" for i in range(max_cols)]
        data_rows = padded
    elif header_rows == 1:
        columns = [
            str(c).strip() if c not in (None, "") else f"Column_{i + 1}"
            for i, c in enumerate(padded[0])
        ]
        data_rows = padded[1:]
    else:
        columns = []
        for i in range(max_cols):
            top = str(padded[0][i]).strip() if padded[0][i] not in (None, "") else ""
            bottom = str(padded[1][i]).strip() if padded[1][i] not in (None, "") else ""
            parts = [p for p in (top, bottom) if p]
            columns.append("_".join(parts) if parts else f"Column_{i + 1}")
        data_rows = padded[2:]

    columns = _dedupe_columns(columns)
    df = pd.DataFrame(data_rows, columns=columns)
    return _coerce_numeric_columns(df)


def _dedupe_columns(columns: list) -> list:
    seen: dict = {}
    result = []
    for col in columns:
        if col not in seen:
            seen[col] = 0
            result.append(col)
        else:
            seen[col] += 1
            result.append(f"{col}_{seen[col]}")
    return result


def _coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """pandas leaves a column with any non-numeric-looking value (e.g. a
    stray "N/A") as dtype=object. If a column is at least 90% numeric once
    that noise is coerced to NaN, treat it as numeric — that's the "mixed
    data types" case explicitly called out for lab exports.
    """
    for col in df.columns:
        non_null = df[col].dropna()
        if non_null.empty or pd.api.types.is_numeric_dtype(df[col]):
            continue
        if non_null.apply(lambda v: isinstance(v, (datetime, pd.Timestamp))).any():
            # pd.to_numeric silently converts datetimes to nanosecond-epoch
            # integers rather than failing — never coerce these, or a real
            # date column turns into "min=1.77e18" garbage.
            continue
        coerced = pd.to_numeric(df[col], errors="coerce")
        if len(coerced.dropna()) / len(non_null) >= 0.9:
            df[col] = coerced
    return df


# ── Report generation (shared by CSV and each Excel sheet) ──────────────────

def _detect_date_columns(df: pd.DataFrame) -> dict:
    date_ranges = {}
    for col in df.columns:
        name_matches = bool(DATE_COLUMN_RE.search(str(col)))
        non_null = df[col].dropna()
        if non_null.empty:
            continue
        if not name_matches and df[col].dtype.kind in "if":
            continue  # don't guess-parse plain numeric columns as dates
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            parsed = df[col].dropna()
        else:
            parsed = pd.to_datetime(df[col], errors="coerce")
        parsed_non_null = parsed.dropna()
        if parsed_non_null.empty:
            continue
        ratio = len(parsed_non_null) / len(non_null)
        if name_matches or ratio >= 0.8:
            date_ranges[str(col)] = (parsed_non_null.min(), parsed_non_null.max())
    return date_ranges


def _detect_lab_value_columns(df: pd.DataFrame, columns: list) -> list:
    lab_cols = []
    for col in columns:
        if LAB_UNIT_NAME_RE.search(col):
            lab_cols.append(col)
            continue
        sample = df[col].dropna().astype(str).head(20)
        if any(LAB_VALUE_RE.search(v) for v in sample):
            lab_cols.append(col)
    return lab_cols


def _column_summary(series: pd.Series, name: str, date_range: Optional[tuple] = None) -> str:
    non_null = series.dropna()
    if non_null.empty:
        return f"{name}: no data"
    if date_range is not None:
        start, end = date_range
        return f"{name}: date range {start.date()} to {end.date()}"
    if pd.api.types.is_numeric_dtype(series):
        return f"{name}: min={non_null.min():.4g}, max={non_null.max():.4g}, mean={non_null.mean():.4g}"
    if pd.api.types.is_datetime64_any_dtype(series):
        return f"{name}: range {non_null.min().date()} to {non_null.max().date()}"
    value_counts = non_null.astype(str).value_counts()
    return f"{name}: {non_null.nunique()} unique values; most common: '{value_counts.index[0]}' ({value_counts.iloc[0]}x)"


def _build_dataframe_report(df: pd.DataFrame, label: str) -> tuple:
    row_count, col_count = df.shape
    columns = [str(c) for c in df.columns]

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    numeric_stats = {}
    for col in numeric_cols:
        series = df[col].dropna()
        if series.empty:
            continue
        numeric_stats[col] = {
            "min": round(float(series.min()), 4),
            "max": round(float(series.max()), 4),
            "mean": round(float(series.mean()), 4),
        }

    date_ranges = _detect_date_columns(df)
    patient_id_cols = [c for c in columns if PATIENT_ID_COLUMN_RE.search(c)]
    date_name_cols = [c for c in columns if DATE_COLUMN_RE.search(c)]
    lab_value_cols = _detect_lab_value_columns(df, columns)

    summary_lines = [
        f"Dataset: {label}",
        f"Rows: {row_count} | Columns: {col_count}",
        f"Columns: {', '.join(columns)}",
    ]
    if numeric_stats:
        summary_lines.append("Numeric columns:")
        for col, s in numeric_stats.items():
            summary_lines.append(f"  - {col}: min={s['min']}, max={s['max']}, mean={s['mean']}")
    if date_ranges:
        parts = [f"{col}: {start.date()} to {end.date()}" for col, (start, end) in date_ranges.items()]
        summary_lines.append("Date range: " + "; ".join(parts))
    else:
        summary_lines.append("Date range: none detected")

    body_lines = []
    if row_count <= _FULL_TABLE_ROW_LIMIT:
        table_rows = [columns] + _rows_as_strings(df)
        body_lines.append(format_table_as_markdown(table_rows))
    else:
        head = df.head(_LARGE_DATASET_PREVIEW_ROWS)
        table_rows = [columns] + _rows_as_strings(head)
        body_lines.append(f"First {_LARGE_DATASET_PREVIEW_ROWS} of {row_count} rows:")
        body_lines.append(format_table_as_markdown(table_rows))
        body_lines.append("")
        body_lines.append("Column summaries (all columns):")
        body_lines.extend(
            f"  - {_column_summary(df[col], col, date_ranges.get(col))}" for col in columns
        )

    report = "\n".join(summary_lines) + "\n\n" + "\n".join(body_lines)

    stats = {
        "row_count": row_count,
        "column_count": col_count,
        "columns": columns,
        "numeric_columns": numeric_stats,
        "date_columns": {col: [str(r[0].date()), str(r[1].date())] for col, r in date_ranges.items()},
        "medical_columns": {
            "patient_id_columns": patient_id_cols,
            "date_columns": date_name_cols,
            "lab_value_columns": lab_value_cols,
        },
    }
    return report, stats


def _rows_as_strings(df: pd.DataFrame) -> list:
    def cell_str(v) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass  # pd.isna can't judge some scalar types — treat as present
        if isinstance(v, (datetime, pd.Timestamp)) and v.time() == datetime.min.time():
            return v.strftime("%Y-%m-%d")
        return str(v)

    return [[cell_str(v) for v in row] for row in df.itertuples(index=False, name=None)]